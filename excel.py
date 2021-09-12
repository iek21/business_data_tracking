from openpyxl import Workbook, load_workbook
from settings import Variables

v = Variables()

class  excel(object):
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "İşletme Verileri"


    def create_excel(self ,data_list):
        self.ws["A1"] = "Tarih :"
        self.ws["A2"] = "Takip Edilen Değer:"

        for i in range(len(data_list)):  ## Raporlma Süresi Geldiğinde kayıt edilen sutun değeri len degeri olarak girilecek
            self.ws.append([data_list[i][0], data_list[i][1]])

        self.wb.save(v.excel_file_name)
